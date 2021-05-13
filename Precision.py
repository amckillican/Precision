# Importing modules
import random
import pygame.gfxdraw
from openpyxl import load_workbook

# Initializing everything
pygame.init()
pygame.font.init()
screen = pygame.display.set_mode((1280, 720))
Title_Font = pygame.font.SysFont("Arial", 60)
Subtitle_Font = pygame.font.SysFont("Arial", 40)
pygame.display.set_caption("Precision - Main Menu")
clock = pygame.time.Clock()
wb = load_workbook(filename="Precision.xlsx")
ws = wb.active
wb.create_sheet("Precision")

# Initialing colors
black = (0, 0, 0)
white = (255, 255, 255)
red = (255, 0, 0)
green = (0, 255, 0)
blue = (0, 0, 255)
background = (39, 41, 44)

# Variables
pos_list_horizontal = [368, 516, 664, 812]
pos_list_vertical = [83, 247, 411, 575]
game_screen = "menu"
game_type = "menu"
game_mode = "start"
average_time = 0
count = 0
num_targets = 0
beginning = 0
reaction_time = 0
start_time = 0
current_time = 0
final_average = 0
game_time = 0
time_left = 60
game_start = 1
start = 1
first = 1
target_count = 4
done = False

# Initialize images
aim_button = pygame.image.load("assets/main_menu/aim_button.png").convert_alpha(screen)
aim_button_s = pygame.image.load("assets/main_menu/aim_button_s.png").convert_alpha(screen)
quit_button = pygame.image.load("assets/main_menu/quit_button.png").convert_alpha(screen)
quit_button_s = pygame.image.load("assets/main_menu/quit_button_s.png").convert_alpha(screen)
reaction_button = pygame.image.load("assets/main_menu/reaction_button.png").convert_alpha(screen)
reaction_button_s = pygame.image.load("assets/main_menu/reaction_button_s.png").convert_alpha(screen)
title = pygame.image.load("assets/main_menu/title.png").convert_alpha(screen)
flick = pygame.image.load("assets/difficulty/flick.png").convert_alpha(screen)
flick_s = pygame.image.load("assets/difficulty/flick_s.png").convert_alpha(screen)
spider_shot = pygame.image.load("assets/difficulty/spider_shot.png").convert_alpha(screen)
spider_shot_s = pygame.image.load("assets/difficulty/spider_shot_s.png").convert_alpha(screen)
grid_shot = pygame.image.load("assets/difficulty/grid_shot.png").convert_alpha(screen)
grid_shot_s = pygame.image.load("assets/difficulty/grid_shot_s.png").convert_alpha(screen)
red_target_image = pygame.transform.scale((pygame.image.load("assets/targets/RED.png").convert_alpha(screen)),
                                          [100, 100])


# Render text function
def Title_text(text="NULL", color=white, position=(640, 360)):
    rendered_text = Title_Font.render(text, True, color)
    rendered_text_rect = rendered_text.get_rect(center=position)
    screen.blit(rendered_text, rendered_text_rect)


# Render text function
def Title_textL(text="NULL", color=white, position=(640, 360)):
    rendered_text = (Title_Font.render(text, True, color))
    rendered_text_rect = rendered_text.get_rect(topleft=position)
    screen.blit(rendered_text, rendered_text_rect)


# Render text function
def Title_textR(text="NULL", color=white, position=(640, 360)):
    rendered_text = (Title_Font.render(text, True, color))
    rendered_text_rect = rendered_text.get_rect(topright=position)
    screen.blit(rendered_text, rendered_text_rect)


# Render text function
def Subtitle_text(text="NULL", color=white, position=(640, 360)):
    rendered_text = (Subtitle_Font.render(text, True, color))
    rendered_text_rect = rendered_text.get_rect(topleft=position)
    screen.blit(rendered_text, rendered_text_rect)


# If the first target is overlapping another, regenerate it
def Regenerate_1():
    global horizontal_red_1, horizontal_red_2, horizontal_red_3, horizontal_red_4, vertical_red_1, vertical_red_2, vertical_red_3, vertical_red_4
    while (
            horizontal_red_1 == horizontal_red_2 or horizontal_red_1 == horizontal_red_3 or horizontal_red_1 == horizontal_red_4) and (
            vertical_red_1 == vertical_red_2 or vertical_red_1 == vertical_red_3 or vertical_red_1 == vertical_red_4):
        horizontal_red_1 = random.choice(pos_list_horizontal)
        vertical_red_1 = random.choice(pos_list_vertical)


# If the second target is overlapping another, regenerate it
def Regenerate_2():
    global horizontal_red_1, horizontal_red_2, horizontal_red_3, horizontal_red_4, vertical_red_1, vertical_red_2, vertical_red_3, vertical_red_4
    while (
            horizontal_red_2 == horizontal_red_1 or horizontal_red_2 == horizontal_red_3 or horizontal_red_2 == horizontal_red_4) and (
            vertical_red_2 == vertical_red_1 or vertical_red_2 == vertical_red_3 or vertical_red_2 == vertical_red_4):
        horizontal_red_2 = random.choice(pos_list_horizontal)
        vertical_red_2 = random.choice(pos_list_vertical)


# If the third target is overlapping another, regenerate it
def Regenerate_3():
    global horizontal_red_1, horizontal_red_2, horizontal_red_3, horizontal_red_4, vertical_red_1, vertical_red_2, vertical_red_3, vertical_red_4
    while (
            horizontal_red_3 == horizontal_red_2 or horizontal_red_3 == horizontal_red_1 or horizontal_red_3 == horizontal_red_4) and (
            vertical_red_3 == vertical_red_2 or vertical_red_3 == vertical_red_1 or vertical_red_3 == vertical_red_4):
        horizontal_red_3 = random.choice(pos_list_horizontal)
        vertical_red_3 = random.choice(pos_list_vertical)


# If the fourth target is overlapping another, regenerate it
def Regenerate_4():
    global horizontal_red_1, horizontal_red_2, horizontal_red_3, horizontal_red_4, vertical_red_1, vertical_red_2, vertical_red_3, vertical_red_4
    while (
            horizontal_red_4 == horizontal_red_2 or horizontal_red_4 == horizontal_red_3 or horizontal_red_4 == horizontal_red_1) and (
            vertical_red_4 == vertical_red_2 or vertical_red_4 == vertical_red_3 or vertical_red_4 == vertical_red_1):
        horizontal_red_4 = random.choice(pos_list_horizontal)
        vertical_red_4 = random.choice(pos_list_vertical)


# Main program loop
while not done:
    # Keep track of the time
    current_time = pygame.time.get_ticks()

    # If the screen is on the start menu
    if game_type == "menu":
        pygame.display.set_caption("Precision - Main Menu")
        # Check the user events
        for event in pygame.event.get():
            # If the close button is pressed
            if event.type == pygame.QUIT:
                # Quit the application
                quit()

            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_ESCAPE:
                    ws['A1'] = None
                    ws['B1'] = None
                    ws['C1'] = None
                    ws['D1'] = None
                    wb.save("Precision.xlsx")

            # Clearing the screen
            screen.fill(background)

            # If the screen is on the main menu
            if game_screen == "menu":
                # Clear the screen
                pygame.Surface.fill(screen, background)

                # Displaying the menu images
                screen.blit(title, [160, 80])
                screen.blit(reaction_button, [750, 262])
                screen.blit(aim_button, [750, 410])
                screen.blit(quit_button, [750, 558])

                # Displaying the high scores
                Subtitle_text("High Scores", white, (190, 275))
                if ws['A1'].value is not None:
                    Subtitle_text(f"Reaction Time: {ws['A1'].value} MS", white, (120, 350))
                elif ws['A1'].value is None:
                    Subtitle_text(f"Reaction Time: {ws['A1'].value}", white, (120, 350))

                if ws['B1'].value is not None:
                    Subtitle_text(f"Flick Shot Time: {ws['B1'].value} MS", white, (120, 425))
                elif ws['B1'].value is None:
                    Subtitle_text(f"Flick Shot Time: {ws['B1'].value}", white, (120, 425))

                if ws['C1'].value is not None:
                    Subtitle_text(f"Spider Shot Time: {ws['C1'].value} MS", white, (120, 500))
                elif ws['C1'].value is None:
                    Subtitle_text(f"Spider Shot Time: {ws['C1'].value}", white, (120, 500))

                if ws['C1'].value is not None:
                    Subtitle_text(f"Grid Shot Time: {ws['D1'].value} MS", white, (120, 575))
                elif ws['C1'].value is None:
                    Subtitle_text(f"Grid Shot Time: {ws['D1'].value}", white, (120, 575))

                # Highlighting the reaction time button if the mouse is overtop of it
                if reaction_button.get_rect(topleft=[750, 262]).collidepoint(pygame.mouse.get_pos()):
                    screen.blit(reaction_button_s, [750, 262])

                    # Start the reaction time game
                    if event.type == pygame.MOUSEBUTTONDOWN and reaction_button.get_rect(
                            topleft=[750, 262]).collidepoint(pygame.mouse.get_pos()):
                        pygame.Surface.fill(screen, background)
                        game_type = "react"

                # Highlighting the aim training button if the mouse is overtop of it
                elif aim_button.get_rect(topleft=[750, 410]).collidepoint(pygame.mouse.get_pos()):
                    screen.blit(aim_button_s, [750, 410])

                    # Change to the aim training selection screen
                    if event.type == pygame.MOUSEBUTTONDOWN and aim_button.get_rect(
                            topleft=[750, 410]).collidepoint(pygame.mouse.get_pos()):
                        pygame.Surface.fill(screen, background)
                        game_screen = "aim"

                # Highlighting the quit button if the mouse is overtop of it
                elif quit_button.get_rect(topleft=[750, 558]).collidepoint(pygame.mouse.get_pos()):
                    screen.blit(quit_button_s, [750, 558])

                    # Quitting the game
                    if event.type == pygame.MOUSEBUTTONDOWN and quit_button.get_rect(
                            topleft=[750, 558]).collidepoint(pygame.mouse.get_pos()):
                        quit()

            # If the menu is on the aim trainer game selection menu
            if game_screen == "aim":
                # Clear the screen
                pygame.Surface.fill(screen, background)

                # Displaying the menu images
                screen.blit(title, [160, 80])
                Title_text(text="Choose A Game", color=white, position=[640, 350])
                screen.blit(flick, [93, 500])
                screen.blit(spider_shot, [389, 500])
                screen.blit(grid_shot, [686, 500])
                screen.blit(quit_button, [983, 500])

                # Highlighting the flicking button if the mouse is overtop of it
                if flick.get_rect(topleft=[93, 500]).collidepoint(pygame.mouse.get_pos()):
                    screen.blit(flick_s, [93, 500])

                    # Change to the flicking game mode
                    if event.type == pygame.MOUSEBUTTONDOWN and flick.get_rect(topleft=[93, 500]).collidepoint(
                            pygame.mouse.get_pos()):
                        game_type = "flick"
                        screen.fill(background)

                # Highlighting the tracking button if the mouse is overtop of it
                elif spider_shot.get_rect(topleft=[389, 500]).collidepoint(pygame.mouse.get_pos()):
                    screen.blit(spider_shot_s, [389, 500])

                    # Change to the tracking game mode
                    if event.type == pygame.MOUSEBUTTONDOWN and spider_shot.get_rect(topleft=[389, 500]).collidepoint(
                            pygame.mouse.get_pos()):
                        game_type = "spider_shot"
                        screen.fill(background)

                # Highlighting the grid shot button if the mouse is overtop of it
                elif grid_shot.get_rect(topleft=[686, 500]).collidepoint(pygame.mouse.get_pos()):
                    screen.blit(grid_shot_s, [686, 500])

                    # Change to the grid shot game mode
                    if event.type == pygame.MOUSEBUTTONDOWN and grid_shot.get_rect(topleft=[686, 500]).collidepoint(
                            pygame.mouse.get_pos()):
                        game_type = "grid_shot"
                        screen.fill(background)

                # Highlighting the flicking button if the mouse is overtop of it
                elif quit_button.get_rect(topleft=[983, 500]).collidepoint(pygame.mouse.get_pos()):
                    screen.blit(quit_button_s, [983, 500])

                    # Quitting the game
                    if event.type == pygame.MOUSEBUTTONDOWN and quit_button.get_rect(
                            topleft=[983, 500]).collidepoint(pygame.mouse.get_pos()):
                        game_screen = "menu"

    # If the reaction game mode is chosen
    if game_type == "react":
        pygame.display.set_caption("Precision - Reaction Time Test")
        screen.fill(background)
        # Tracking events
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                quit()

            # If key is pressed
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_ESCAPE:
                    game_screen = "menu"
                    game_type = "menu"
                    game_mode = "start"
                    average_time = 0
                    count = 0
                    reaction_time = 0
                    start_time = 0
                    current_time = 0

                if game_mode == "start":
                    game_mode = "wait"

                    # Run the game 3 times in a row
                    if count >= 1:
                        Title_text(f"Reaction Time: {reaction_time}", white, (640, 600))

                    start_time = current_time + random.randint(1000, 4000)

                # Measure the reaction time
                if game_mode == "wait_for_reaction":
                    game_mode = "wait"
                    reaction_time = (current_time - start_time) / 1000
                    start_time = current_time + random.randint(1000, 4000)
                    count += 1
                    average_time = ((average_time * (count - 1) + reaction_time) / count)

            # If the results are being displayed
            if game_mode == "results":
                # If the user presses the enter key
                if event.type == pygame.KEYDOWN:
                    if event.key == pygame.K_RETURN:
                        # Return to the main menu
                        if ws['A1'].value is None:
                            ws['A1'] = int(f"{final_average:.0f}")
                        if ws['A1'] is not None and final_average <= ws['A1'].value:
                            ws['A1'] = int(f"{final_average:.0f}")
                        wb.save("Precision.xlsx")
                        game_screen = "menu"
                        game_type = "menu"
                        game_mode = "start"
                        average_time = 0
                        count = 0
                        reaction_time = 0
                        start_time = 0
                        current_time = 0

        # Showing the previous reaction time
        if game_mode == "wait":
            if count >= 1:
                Title_text(f"Reaction Time: {reaction_time * 1000:.0f} MS", white, (640, 600))

            if current_time >= start_time:
                game_mode = "wait_for_reaction"

                # Clearing the screen
                screen.fill(background)

        # After 3 tries display the results
        if count == 3:
            game_mode = "results"

        # Prompt the user to start
        if count < 3:
            if game_mode == "start":
                Title_text("Press Any Key To Start")
            if game_mode == "wait_for_reaction":
                Title_text("Press Any Key")

                if count >= 1:
                    Title_text(f"Reaction Time: {reaction_time * 1000:.0f} MS", white, (640, 600))

        # Saving the users name and reaction time
        if game_mode == "results":
            screen.fill(background)
            final_average = average_time * 1000
            Title_text("Average Reaction", white, (640, 250))
            Title_text(f"Time: {final_average:.0f} MS", white, (640, 320))
            Title_text("Press Enter To Exit To Menu", white, (640, 450))

    # If the flicking game mode is chosen
    if game_type == "flick":
        pygame.display.set_caption("Precision - Flick Training")
        # Clear the screen if it is the first time the code is ran
        if beginning == 0:
            screen.fill(background)
            beginning = 1

        # Tracking user events
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                quit()

            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_ESCAPE:
                    game_type = "menu"
                    game_screen = "menu"
                    game_mode = "start"
                    num_targets = 0
                    beginning = 0
                    start_time = 0
                    game_time = 0
                    time_left = 0

            # If the game is waiting to start
            if game_mode == "start":
                # If the user clicks the mouse
                if event.type == pygame.MOUSEBUTTONDOWN:
                    # Start a timer and the game
                    screen.fill(background)
                    start_time = current_time
                    game_mode = "generate"

            # While the game has been running for less than 60 seconds
            if game_mode != "results" and game_mode != "start":
                if game_time < 60:
                    # Generate a target at a random position
                    if game_mode == "generate":
                        horizontal_red = random.randint(0, 1020)
                        vertical_red = random.randint(75, 520)
                        screen.blit(red_target_image, [horizontal_red, vertical_red])
                        game_mode = "react"

                    # Wait for the user to click the target and increase the counter when they do
                    if game_mode == "react":
                        if event.type == pygame.MOUSEBUTTONDOWN and red_target_image.get_rect(
                                topleft=[horizontal_red, vertical_red]).collidepoint(pygame.mouse.get_pos()):
                            screen.fill(background)
                            num_targets += 1
                            game_mode = "generate"

            # If the results are being displayed
            if game_mode == "results":
                # If the user presses the enter key
                if event.type == pygame.KEYDOWN:
                    if event.key == pygame.K_RETURN:
                        # Return to the main menu
                        if ws['B1'].value is None:
                            ws['B1'] = int(f"{final_average:.0f}")
                        if ws['B1'] is not None and final_average <= ws['B1'].value:
                            ws['B1'] = int(f"{final_average:.0f}")
                        wb.save("Precision.xlsx")
                        game_type = "menu"
                        game_screen = "menu"
                        game_mode = "start"
                        num_targets = 0
                        beginning = 0
                        start_time = 0
                        game_time = 0
                        time_left = 0

        # While the game has been running for more than 60 seconds
        if game_mode != "results" and game_mode != "start":
            if game_time > 60:
                # Display the results
                game_mode = "results"
                screen.fill(background)

            game_time = (current_time - start_time) / 1000

            # Calculating how much time is left
            time_left = (60 - game_time) // 1

            # Show the targets hit and timer
            pygame.draw.rect(screen, (10, 20, 20), (0, 0, 1280, 75))
            Title_textL("Targets: " + str(num_targets), white, (10, 0))
            Title_textR(str(int(time_left)), white, (1270, 0))

        # If the game has not started
        if game_mode == "start":
            # Display the starting text
            Title_text("Click To Start")

        # If the time has run out
        if game_mode == "results":
            # Display the results
            screen.fill(background)
            if num_targets == 0:
                num_targets += 1
            final_average = 60000 / num_targets
            Title_text(f"Targets Hit: {num_targets}", white, (640, 100))
            Title_text(f"Average Response", white, (640, 250))
            Title_text(f"Time: {final_average:.0f} MS", white, (640, 320))
            Title_text("Press Enter To Exit To Menu", white, (640, 500))

    # If the spider shot game mode is chosen
    if game_type == "spider_shot":
        pygame.display.set_caption("Precision - Spider Shot")
        # Clear the screen if it is the first time the code is ran
        if beginning == 0:
            screen.fill(background)
            beginning = 1

        # Tracking user events
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                quit()

            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_ESCAPE:
                    game_type = "menu"
                    game_screen = "menu"
                    game_mode = "start"
                    num_targets = 0
                    beginning = 0
                    start_time = 0
                    game_time = 0
                    time_left = 0

            # If the game is waiting to start
            if game_mode == "start":
                # If the user clicks the mouse
                if event.type == pygame.MOUSEBUTTONDOWN:
                    # Start a timer and the game
                    screen.fill(background)
                    start_time = current_time
                    game_mode = "generate"

            # While the game has been running for less than 60 seconds
            if game_mode != "results" and game_mode != "start":
                if game_time < 60:
                    # Generate a target at a random position
                    if game_mode == "generate":
                        if (num_targets + 2) % 2 == 0:
                            screen.fill(background)
                            horizontal_red = 590
                            vertical_red = 310
                            screen.blit(red_target_image, [horizontal_red, vertical_red])
                            game_mode = "react"

                        elif (num_targets + 2) % 2 != 0:
                            screen.fill(background)
                            horizontal_red = random.randint(0, 1020)
                            vertical_red = random.randint(75, 520)
                            screen.blit(red_target_image, [horizontal_red, vertical_red])
                            game_mode = "react"

                    # Wait for the user to click the target and increase the counter when they do
                    if game_mode == "react":
                        if event.type == pygame.MOUSEBUTTONDOWN and red_target_image.get_rect(
                                topleft=[horizontal_red, vertical_red]).collidepoint(pygame.mouse.get_pos()):
                            num_targets += 1
                            start = 1
                            first = 1
                            game_mode = "generate"

            # If the results are being displayed
            if game_mode == "results":
                # If the user presses the enter key
                if event.type == pygame.KEYDOWN:
                    if event.key == pygame.K_RETURN:
                        # Return to the main menu
                        if ws['C1'].value is None:
                            ws['C1'] = int(f"{final_average:.0f}")
                        if ws['C1'] is not None and final_average <= ws['C1'].value:
                            ws['C1'] = int(f"{final_average:.0f}")
                        wb.save("Precision.xlsx")
                        game_type = "menu"
                        game_screen = "menu"
                        game_mode = "start"
                        num_targets = 0
                        beginning = 0
                        start_time = 0
                        game_time = 0
                        time_left = 0

        # Wait for the user to click the target and increase the counter when they do
        if game_mode == "react":
            # Moving the target
            if (num_targets + 2) % 2 != 0:
                screen.fill(background)
                if start == 1:
                    target_speed_x = 7 * random.choice((1, -1))
                    target_speed_y = 7 * random.choice((1, -1))
                    start = 0

                horizontal_red += target_speed_x
                vertical_red += target_speed_y
                screen.blit(red_target_image, [horizontal_red, vertical_red])

                if vertical_red <= 80 or vertical_red >= 620:
                    target_speed_y *= -1
                if horizontal_red <= 0 or horizontal_red >= 1180:
                    target_speed_x *= -1

        # If the game has not started
        if game_mode == "start":
            # Display the starting text
            Title_text("Click To Start")

        # If the game has not started yet and has not finished
        if game_mode != "results" and game_mode != "start":
            # Show the targets hit and timer
            pygame.draw.rect(screen, (10, 20, 20), (0, 0, 1280, 75))
            Title_textL("Targets: " + str(num_targets), white, (10, 0))
            Title_textR(str(int(time_left)), white, (1270, 0))

            # Setting the game timer
            game_time = (current_time - start_time) / 1000

            # Calculating how much time is left
            time_left = (60 - game_time) // 1

            if game_time > 60:
                # Display the results
                game_mode = "results"
                screen.fill(background)

        # If the time has run out
        if game_mode == "results":
            # Display the results
            screen.fill(background)
            final_average = 60000 / num_targets
            Title_text(f"Targets Hit: {num_targets}", white, (640, 100))
            Title_text(f"Average Response", white, (640, 250))
            Title_text(f"Time: {final_average:.0f} MS", white, (640, 320))
            Title_text("Press Enter To Exit To Menu", white, (640, 500))

    # If the grid shot game mode is chosen
    if game_type == "grid_shot":
        pygame.display.set_caption("Precision - Grid Shot")
        # Clear the screen if it is the first time the code is ran
        if beginning == 0:
            screen.fill(background)
            beginning = 1

        # Setting the game timer
        if game_mode != "results" and game_mode != "start":
            game_time = (current_time - start_time) / 1000

            # Calculating how much time is left
            time_left = (60 - game_time) // 1

        # Tracking user events
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                quit()

            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_ESCAPE:
                    game_type = "menu"
                    game_screen = "menu"
                    game_mode = "start"
                    num_targets = 0
                    beginning = 0
                    start_time = 0
                    game_time = 0
                    time_left = 0

            # If the game is waiting to start
            if game_mode == "start":
                # If the user clicks the mouse
                if event.type == pygame.MOUSEBUTTONDOWN:
                    # Start a timer and the game
                    screen.fill(background)
                    start_time = current_time
                    game_mode = "react"

                    if game_start == 1:
                        # Generate the targets at a random set position
                        horizontal_red_1 = random.choice(pos_list_horizontal)
                        vertical_red_1 = random.choice(pos_list_vertical)

                        horizontal_red_2 = random.choice(pos_list_horizontal)
                        vertical_red_2 = random.choice(pos_list_vertical)

                        horizontal_red_3 = random.choice(pos_list_horizontal)
                        vertical_red_3 = random.choice(pos_list_vertical)

                        horizontal_red_4 = random.choice(pos_list_horizontal)
                        vertical_red_4 = random.choice(pos_list_vertical)

                        # Regenerate them if they are in the same spot
                        Regenerate_1()
                        Regenerate_2()
                        Regenerate_3()
                        Regenerate_4()

                        # Display them on screen
                        screen.blit(red_target_image, [horizontal_red_1, vertical_red_1])
                        screen.blit(red_target_image, [horizontal_red_2, vertical_red_2])
                        screen.blit(red_target_image, [horizontal_red_3, vertical_red_3])
                        screen.blit(red_target_image, [horizontal_red_4, vertical_red_4])

            # Wait for the user to click the target and increase the counter when they do
            if game_mode == "react":
                if event.type == pygame.MOUSEBUTTONDOWN and red_target_image.get_rect(
                        topleft=[horizontal_red_1, vertical_red_1]).collidepoint(pygame.mouse.get_pos()):
                    pygame.draw.rect(screen, background, (horizontal_red_1, vertical_red_1, 100, 100))
                    num_targets += 1
                    target_count -= 1
                    game_mode = "generate_1"

                # Wait for the user to click the target and increase the counter when they do
                elif event.type == pygame.MOUSEBUTTONDOWN and red_target_image.get_rect(
                        topleft=[horizontal_red_2, vertical_red_2]).collidepoint(pygame.mouse.get_pos()):
                    pygame.draw.rect(screen, background, (horizontal_red_2, vertical_red_2, 100, 100))
                    num_targets += 1
                    target_count -= 1
                    game_mode = "generate_2"

                # Wait for the user to click the target and increase the counter when they do
                elif event.type == pygame.MOUSEBUTTONDOWN and red_target_image.get_rect(
                        topleft=[horizontal_red_3, vertical_red_3]).collidepoint(pygame.mouse.get_pos()):
                    pygame.draw.rect(screen, background, (horizontal_red_3, vertical_red_3, 100, 100))
                    num_targets += 1
                    target_count -= 1
                    game_mode = "generate_3"

                # Wait for the user to click the target and increase the counter when they do
                elif event.type == pygame.MOUSEBUTTONDOWN and red_target_image.get_rect(
                        topleft=[horizontal_red_4, vertical_red_4]).collidepoint(pygame.mouse.get_pos()):
                    pygame.draw.rect(screen, background, (horizontal_red_4, vertical_red_4, 100, 100))
                    num_targets += 1
                    target_count -= 1
                    game_mode = "generate_4"

            # If the results are being displayed
            if game_mode == "results":
                # If the user presses the enter key
                if event.type == pygame.KEYDOWN:
                    if event.key == pygame.K_RETURN:
                        # Return to the main menu
                        if ws['D1'].value is None:
                            ws['D1'] = int(f"{final_average:.0f}")
                        if ws['D1'] is not None and final_average <= ws['D1'].value:
                            ws['D1'] = int(f"{final_average:.0f}")
                        wb.save("Precision.xlsx")
                        game_type = "menu"
                        game_screen = "menu"
                        game_mode = "start"
                        num_targets = 0
                        beginning = 0
                        start_time = 0
                        game_time = 0
                        time_left = 0

        # While the game has been running for less than 60 seconds
        if game_mode != "results" and game_mode != "start":
            # Generate the first target
            if game_mode == "generate_1":
                if game_time < 60:
                    # Generate a target at a random position
                    horizontal_red_1 = random.choice(pos_list_horizontal)
                    vertical_red_1 = random.choice(pos_list_vertical)
                    Regenerate_1()
                    screen.blit(red_target_image, [horizontal_red_1, vertical_red_1])
                    target_count += 1
                    game_mode = "react"

            # Generate the second target
            elif game_mode == "generate_2":
                if game_time < 60:
                    # Generate a target at a random position
                    horizontal_red_2 = random.choice(pos_list_horizontal)
                    vertical_red_2 = random.choice(pos_list_vertical)
                    Regenerate_2()
                    screen.blit(red_target_image, [horizontal_red_2, vertical_red_2])
                    target_count += 1
                    game_mode = "react"

            # Generate the third target
            elif game_mode == "generate_3":
                if game_time < 60:
                    # Generate a target at a random position
                    horizontal_red_3 = random.choice(pos_list_horizontal)
                    vertical_red_3 = random.choice(pos_list_vertical)
                    Regenerate_3()
                    screen.blit(red_target_image, [horizontal_red_3, vertical_red_3])
                    target_count += 1
                    game_mode = "react"

            # Generate the fourth target
            elif game_mode == "generate_4":
                if game_time < 60:
                    # Generate a target at a random position
                    horizontal_red_4 = random.choice(pos_list_horizontal)
                    vertical_red_4 = random.choice(pos_list_vertical)
                    Regenerate_4()
                    screen.blit(red_target_image, [horizontal_red_4, vertical_red_4])
                    target_count += 1
                    game_mode = "react"

        # If the game has not started
        if game_mode == "start":
            # Display the starting text
            Title_text("Click To Start")

        # If the game has not started yet and has not finished
        if game_mode != "results" and game_mode != "start":
            # Show the targets hit and timer
            pygame.draw.rect(screen, (10, 20, 20), (0, 0, 1280, 75))
            Title_textL("Targets: " + str(num_targets), white, (10, 0))
            Title_textR(str(int(time_left)), white, (1270, 0))

        # While the game has been running for more than 60 seconds
        if game_mode != "results" and game_mode != "start":
            if game_time > 60:
                # Display the results
                game_mode = "results"
                screen.fill(background)

        # If the time has run out
        if game_mode == "results":
            # Display the results
            screen.fill(background)
            final_average = 60000 / num_targets
            Title_text(f"Targets Hit: {num_targets}", white, (640, 100))
            Title_text(f"Average Response", white, (640, 250))
            Title_text(f"Time: {final_average:.0f} MS", white, (640, 320))
            Title_text("Press Enter To Exit To Menu", white, (640, 500))

    # Update the screen with what we've drawn
    pygame.display.flip()

    # Limit to 60 fps
    clock.tick(60)

# Close the window and quit
pygame.quit()
