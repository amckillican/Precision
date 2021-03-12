import random

import pygame.gfxdraw
from openpyxl import Workbook
from openpyxl import load_workbook

# Initializing everything
pygame.init()
pygame.font.init()
Title_Font = pygame.font.SysFont("Arial", 60)
Subtitle_font = pygame.font.SysFont("Arial", 30)

# Variables
done = False
game_state = "start"
count = 0
start_time = 0
average_time = 0
name = ""
split_name = []
reaction_time = 0
current_time = pygame.time.get_ticks()

# Spreadsheet stuff
wb = Workbook()
# noinspection PyRedeclaration
wb = load_workbook("Precision.xlsx")
ws = wb.active


font = pygame.font.SysFont("Calibri", 30)

# Initialing colors
black = (0, 0, 0)
white = (255, 255, 255)
red = (255, 0, 0)
green = (0, 255, 0)
blue = (0, 0, 255)
background = (39, 41, 44)

# Set the width and height of the screen [width, height]
res = (1280, 720)
screen = pygame.display.set_mode(res)


# Render some text
def Title_text(text="NULL", color=white, position=(640, 360)):
    rendered_text = (Title_Font.render(text, True, color))
    rendered_text_rect = rendered_text.get_rect(center=position)
    screen.blit(rendered_text, rendered_text_rect)


# Render some text
def Subtitle_text(text="NULL", color=white, position=(640, 360)):
    rendered_text = (Subtitle_font.render(text, True, color))
    rendered_text_rect = rendered_text.get_rect(center=position)
    screen.blit(rendered_text, rendered_text_rect)


def Reset_and_save():
    global name, final_average, game_state, count, average_time, reaction_time
    ws["A1"] = name
    ws["B1"] = f"{final_average:.0f} MS"
    ws.insert_rows(1)
    wb.save("Precision.xlsx")
    game_state = "start"
    count = 0
    average_time = 0
    reaction_time = 0
    name = ""


# Setting the title of the window
pygame.display.set_caption("Reaction Time Test")

# Setting the screen refresh rate
fps = 60
clock = pygame.time.Clock()

# Main program loop
while not done:
    # Main event loop
    screen.fill(background)
    current_time = pygame.time.get_ticks()

    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False
            pygame.quit()

        if event.type == pygame.KEYDOWN:
            if event.key == pygame.K_ESCAPE:
                ws.delete_cols(1, 2)
                wb.save("Precision.xlsx")

        if event.type == pygame.KEYDOWN and event.key != pygame.K_ESCAPE:
            if game_state == "start":
                game_state = "wait"

                if count >= 1:
                    Title_text(f"Reaction Time: {reaction_time}", white, (640, 600))

                start_time = current_time + random.randint(1000, 4000)
            if game_state == "wait_for_reaction":
                game_state = "wait"
                reaction_time = (current_time - start_time) / 1000
                start_time = current_time + random.randint(1000, 4000)
                count += 1
                average_time = ((average_time * (count - 1) + reaction_time) / count)

        if game_state == "results":
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_a:
                    name += str(chr(event.key))
                if event.key == pygame.K_b:
                    name += str(chr(event.key))
                if event.key == pygame.K_c:
                    name += str(chr(event.key))
                if event.key == pygame.K_d:
                    name += str(chr(event.key))
                if event.key == pygame.K_e:
                    name += str(chr(event.key))
                if event.key == pygame.K_f:
                    name += str(chr(event.key))
                if event.key == pygame.K_g:
                    name += str(chr(event.key))
                if event.key == pygame.K_h:
                    name += str(chr(event.key))
                if event.key == pygame.K_i:
                    name += str(chr(event.key))
                if event.key == pygame.K_j:
                    name += str(chr(event.key))
                if event.key == pygame.K_k:
                    name += str(chr(event.key))
                if event.key == pygame.K_l:
                    name += str(chr(event.key))
                if event.key == pygame.K_m:
                    name += str(chr(event.key))
                if event.key == pygame.K_n:
                    name += str(chr(event.key))
                if event.key == pygame.K_o:
                    name += str(chr(event.key))
                if event.key == pygame.K_p:
                    name += str(chr(event.key))
                if event.key == pygame.K_q:
                    name += str(chr(event.key))
                if event.key == pygame.K_r:
                    name += str(chr(event.key))
                if event.key == pygame.K_s:
                    name += str(chr(event.key))
                if event.key == pygame.K_t:
                    name += str(chr(event.key))
                if event.key == pygame.K_u:
                    name += str(chr(event.key))
                if event.key == pygame.K_v:
                    name += str(chr(event.key))
                if event.key == pygame.K_w:
                    name += str(chr(event.key))
                if event.key == pygame.K_x:
                    name += str(chr(event.key))
                if event.key == pygame.K_y:
                    name += str(chr(event.key))
                if event.key == pygame.K_z:
                    name += str(chr(event.key))
                if event.key == pygame.K_SPACE:
                    name += str(chr(event.key))
                if event.key == pygame.K_BACKSPACE:
                    name = name[:-1]
                if event.key == pygame.K_RETURN:
                    Reset_and_save()

    if game_state == "wait":
        if count >= 1:
            Title_text(f"Reaction Time: {reaction_time*1000:.0f} MS", white, (640, 600))

        if current_time >= start_time:
            game_state = "wait_for_reaction"

            # Clearing the screen
            screen.fill(background)


    if count == 3:
        game_state = "results"

    if count < 3:
        center = screen.get_rect().center
        if game_state == "start":
            Title_text("Press Any Key To Start")
        if game_state == "wait_for_reaction":
            Title_text("Press Any Key", white)

            if count >= 1:
                Title_text(f"Reaction Time: {reaction_time*1000:.0f} MS", white, (640, 600))

    if game_state == "results":
        screen.fill(background)
        final_average = average_time * 1000
        Title_text("Average Reaction", white, (640, 100))
        Title_text(f"Time: {final_average:.0f} MS", white, (640, 200))
        Title_text("Please Type Your", white, (640, 450))
        Title_text(f"Name: {name}", white, (640, 550))

    # Update the screen with what we've drawn
    pygame.display.flip()

    # Limit to 60 fps
    clock.tick(fps)

# Close the window and quit
pygame.quit()
